from __future__ import annotations

from io import BytesIO
import re
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill

from explainer import generate_whatsapp_message


EXPORT_COLUMNS = [
    "Rank",
    "Name",
    "Score",
    "Recommendation",
    "Matched Skills",
    "Missing Skills",
    "Experience (Yrs)",
    "Notice Period",
    "Current CTC",
    "Expected CTC",
    "Location",
    "Current Company",
    "Email",
    "Phone",
    "Explanation",
]


def _to_export_rows(candidates: List[Dict]) -> List[Dict]:
    rows = []
    for c in candidates:
        rows.append(
            {
                "Rank": c.get("rank"),
                "Name": c.get("name"),
                "Score": c.get("total_score"),
                "Recommendation": c.get("recommendation"),
                "Matched Skills": ", ".join(c.get("matched_required", [])),
                "Missing Skills": ", ".join(c.get("missing_skills", [])),
                "Experience (Yrs)": c.get("total_experience_years"),
                "Notice Period": c.get("notice_period"),
                "Current CTC": c.get("current_ctc"),
                "Expected CTC": c.get("expected_ctc"),
                "Location": c.get("location"),
                "Current Company": c.get("current_company"),
                "Email": c.get("email"),
                "Phone": c.get("phone"),
                "Explanation": c.get("explanation"),
            }
        )
    return rows


def export_to_excel(candidates: List[Dict], job_title: str) -> bytes:
    export_rows = _to_export_rows(candidates)
    # Keep a stable schema even for empty candidate lists (e.g., empty shortlist).
    df = pd.DataFrame(export_rows, columns=EXPORT_COLUMNS)

    output = BytesIO()
    base_sheet_name = job_title or "Resume_Screening"
    safe_sheet_name = re.sub(r"[\\/*?:\[\]]", "_", base_sheet_name).strip()
    sheet_name = (safe_sheet_name or "Resume_Screening")[:31]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        high_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        mid_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        low_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        if "Score" in df.columns:
            score_col_idx = list(df.columns).index("Score") + 1

            for row_idx in range(2, ws.max_row + 1):
                score_value = ws.cell(row=row_idx, column=score_col_idx).value or 0
                try:
                    score_value = float(score_value)
                except (TypeError, ValueError):
                    score_value = 0

                if score_value >= 80:
                    fill = high_fill
                elif score_value >= 60:
                    fill = mid_fill
                else:
                    fill = low_fill

                for cell in ws[row_idx]:
                    cell.fill = fill

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()


def generate_candidate_messages(candidates: List[Dict], jd_data: Dict, api_key: str | None = None) -> Dict[str, str]:
    messages = {}
    for candidate in candidates:
        candidate_name = candidate.get("name") or "Candidate"
        messages[candidate_name] = generate_whatsapp_message(candidate, jd_data, api_key=api_key)
    return messages
